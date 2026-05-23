from __future__ import annotations

import csv
import io

from fastapi import APIRouter, File, HTTPException, Request, Response, UploadFile, status
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from backend.dependencies import CurrentAdmin, DbSession
from backend.grade_utils import format_grade_label
from backend.models.audit_log import AuditLog
from backend.models.conversation import Conversation, Message
from backend.models.schemas import (
    AuditLogRead,
    ClassroomOptionRead,
    ConversationArchiveMessageRead,
    ConversationArchiveRead,
    PasswordResetRequest,
    UserCreate,
    UserImportIssue,
    UserImportResult,
    UserRead,
    UserUpdate,
)
from backend.models.user import Classroom, User, UserRole
from backend.security import get_password_hash
from backend.services.account_service import build_default_password, build_generated_username
from backend.services.audit_service import audit_service
from backend.services.auth_service import auth_service
from backend.services.student_grade_service import student_grade_service

router = APIRouter(prefix="/api/admin", tags=["admin"])
MANAGED_USER_ROLES = {UserRole.STUDENT, UserRole.TEACHER}


def _normalized_classroom_name(value: str | None) -> str | None:
    if value is None:
        return None
    normalized = value.strip()
    return normalized or None


def _parse_managed_role(value: str | UserRole | None) -> UserRole | None:
    if isinstance(value, UserRole):
        return value if value in MANAGED_USER_ROLES else None
    normalized = str(value or "").strip().lower()
    mapping = {
        "student": UserRole.STUDENT,
        "学生": UserRole.STUDENT,
        "teacher": UserRole.TEACHER,
        "教师": UserRole.TEACHER,
    }
    return mapping.get(normalized)


def _validate_managed_user_fields(role: UserRole, *, grade: int | None, classroom_name: str | None, is_graduated: bool) -> None:
    if role not in MANAGED_USER_ROLES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only student and teacher accounts can be managed here")

    if role == UserRole.STUDENT:
        if grade not in {1, 2, 3} and not is_graduated:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Student grade must be 高一/高二/高三 or 毕业")
        if not classroom_name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Student classroom is required")
        return

    if grade is not None or classroom_name is not None or is_graduated:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Teacher accounts do not use grade/classroom fields")


def _get_or_create_classroom(db: DbSession, grade: int | None, classroom_name: str | None) -> Classroom | None:
    if grade is None or not classroom_name:
        return None
    classroom = db.scalar(select(Classroom).where(Classroom.grade == grade, Classroom.name == classroom_name))
    if classroom:
        return classroom
    classroom = Classroom(grade=grade, name=classroom_name)
    db.add(classroom)
    db.commit()
    db.refresh(classroom)
    return classroom


def _ensure_generated_username_available(db: DbSession, username: str, *, exclude_user_id: int | None = None) -> None:
    statement = select(User).where(User.username == username)
    if exclude_user_id is not None:
        statement = statement.where(User.id != exclude_user_id)
    if db.scalar(statement.limit(1)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Generated username already exists: {username}")


def _apply_managed_user_payload(
    db: DbSession,
    user: User,
    *,
    full_name: str,
    role: UserRole,
    grade: int | None,
    classroom_name: str | None,
    is_graduated: bool,
    is_active: bool | None = None,
    exclude_user_id: int | None = None,
) -> None:
    normalized_name = full_name.strip()
    if not normalized_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Full name is required")
    normalized_classroom_name = _normalized_classroom_name(classroom_name)
    _validate_managed_user_fields(
        role,
        grade=grade,
        classroom_name=normalized_classroom_name,
        is_graduated=is_graduated,
    )

    if role == UserRole.STUDENT:
        classroom = _get_or_create_classroom(db, grade, normalized_classroom_name)
        username = build_generated_username(normalized_name, role, normalized_classroom_name)
        student_grade_service.apply_manual_grade_state(user, grade=grade, is_graduated=is_graduated)
        user.classroom_id = classroom.id if classroom else None
        user.must_change_password = True
    else:
        username = build_generated_username(normalized_name, role)
        user.grade = None
        user.graduated_at = None
        user.last_grade_promotion_year = None
        user.classroom_id = None
        user.must_change_password = False

    _ensure_generated_username_available(db, username, exclude_user_id=exclude_user_id)
    user.username = username
    user.student_no = None
    user.full_name = normalized_name
    user.role = role
    if is_active is not None:
        user.is_active = is_active


@router.get("/users", response_model=list[UserRead])
def list_users(
    db: DbSession,
    current_user: CurrentAdmin,
    grade: str | None = None,
    classroom_name: str | None = None,
    keyword: str | None = None,
) -> list[UserRead]:
    statement = select(User).options(selectinload(User.classroom)).order_by(User.created_at.desc())

    normalized_grade = (grade or "").strip().lower()
    if normalized_grade in {"1", "2", "3"}:
        statement = statement.where(User.grade == int(normalized_grade), User.graduated_at.is_(None))
    elif normalized_grade == "unset":
        statement = statement.where(User.grade.is_(None), User.graduated_at.is_(None))
    elif normalized_grade == "graduated":
        statement = statement.where(User.graduated_at.is_not(None))

    normalized_classroom_name = _normalized_classroom_name(classroom_name)
    if normalized_classroom_name:
        statement = statement.join(User.classroom).where(Classroom.name.contains(normalized_classroom_name))

    normalized_keyword = (keyword or "").strip()
    if normalized_keyword:
        statement = statement.where(User.full_name.contains(normalized_keyword))

    users = db.scalars(statement).all()
    return [UserRead.model_validate(user) for user in users]


@router.get("/classrooms", response_model=list[ClassroomOptionRead])
def list_classrooms(
    db: DbSession,
    current_user: CurrentAdmin,
    grade: str | None = None,
) -> list[ClassroomOptionRead]:
    statement = select(Classroom).order_by(Classroom.grade.asc(), Classroom.name.asc())
    normalized_grade = (grade or "").strip().lower()
    if normalized_grade in {"1", "2", "3"}:
        statement = statement.where(Classroom.grade == int(normalized_grade))
    elif normalized_grade in {"unset", "graduated"}:
        return []

    classrooms = db.scalars(statement).all()
    return [
        ClassroomOptionRead(
            id=classroom.id,
            grade=classroom.grade,
            name=classroom.name,
            label=f"{format_grade_label(classroom.grade) or ''}{classroom.name}",
        )
        for classroom in classrooms
    ]


@router.post("/users", response_model=UserRead, status_code=status.HTTP_201_CREATED)
def create_user(payload: UserCreate, db: DbSession, current_user: CurrentAdmin, request: Request) -> UserRead:
    role = _parse_managed_role(payload.role)
    if role is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only student and teacher accounts can be created here")

    user = User(
        username="pending",
        student_no=None,
        full_name=payload.full_name.strip(),
        role=role,
        password_hash=get_password_hash(build_default_password(payload.full_name)),
        classroom_id=None,
        must_change_password=role == UserRole.STUDENT,
    )
    _apply_managed_user_payload(
        db,
        user,
        full_name=payload.full_name,
        role=role,
        grade=payload.grade,
        classroom_name=payload.classroom_name,
        is_graduated=payload.is_graduated,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    audit_service.log(
        db,
        actor=current_user,
        action="create_user",
        target_type="user",
        target_id=str(user.id),
        result="success",
        ip_address=request.client.host if request.client else None,
        detail={"role": user.role.value, "username": user.username},
    )
    return UserRead.model_validate(user)


@router.put("/users/{user_id}", response_model=UserRead)
def update_user(
    user_id: int,
    payload: UserUpdate,
    db: DbSession,
    current_user: CurrentAdmin,
    request: Request,
) -> UserRead:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.role == UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Admin account cannot be edited here")

    role = _parse_managed_role(payload.role)
    if role is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only student and teacher accounts can be edited here")

    _apply_managed_user_payload(
        db,
        user,
        full_name=payload.full_name,
        role=role,
        grade=payload.grade,
        classroom_name=payload.classroom_name,
        is_graduated=payload.is_graduated,
        is_active=payload.is_active,
        exclude_user_id=user_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    audit_service.log(
        db,
        actor=current_user,
        action="update_user",
        target_type="user",
        target_id=str(user.id),
        result="success",
        ip_address=request.client.host if request.client else None,
        detail={
            "username": user.username,
            "role": user.role.value,
            "grade": user.grade,
            "classroom_name": user.classroom_name,
            "is_active": user.is_active,
        },
    )
    return UserRead.model_validate(user)


@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: DbSession, current_user: CurrentAdmin, request: Request) -> dict[str, str]:
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cannot delete current admin account")
    if user.role == UserRole.ADMIN:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Admin account cannot be deleted here")
    if user.role == UserRole.STUDENT:
        has_archived_conversations = db.scalar(select(Conversation.id).where(Conversation.student_id == user.id).limit(1))
        if has_archived_conversations is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Student with archived conversations cannot be deleted; deactivate the account instead",
            )

    username = user.username
    db.delete(user)
    db.commit()
    audit_service.log(
        db,
        actor=current_user,
        action="delete_user",
        target_type="user",
        target_id=str(user_id),
        result="success",
        ip_address=request.client.host if request.client else None,
        detail={"username": username},
    )
    return {"status": "ok"}


@router.post("/users/reset-password", response_model=UserRead)
def reset_password(payload: PasswordResetRequest, db: DbSession, current_user: CurrentAdmin, request: Request) -> UserRead:
    user = db.get(User, payload.user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    updated = auth_service.update_password(db, user, build_default_password(user.full_name))
    updated.must_change_password = updated.role == UserRole.STUDENT
    db.add(updated)
    db.commit()
    db.refresh(updated)
    audit_service.log(
        db,
        actor=current_user,
        action="reset_password",
        target_type="user",
        target_id=str(updated.id),
        result="success",
        ip_address=request.client.host if request.client else None,
        detail={"username": updated.username},
    )
    return UserRead.model_validate(updated)


def _parse_import_rows(file_name: str, content: bytes) -> list[dict[str, str]]:
    suffix = (file_name or "").lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        return [dict(row) for row in reader]
    if suffix == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="缺少 openpyxl，当前无法解析 xlsx") from exc

        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
        return [{headers[index]: str(value).strip() if value is not None else "" for index, value in enumerate(row)} for row in sheet.iter_rows(min_row=2, values_only=True)]
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only csv/xlsx supported")


def _import_users(file_name: str, content: bytes, db: DbSession, current_user: CurrentAdmin, request: Request | None) -> UserImportResult:
    rows = _parse_import_rows(file_name, content)
    created = 0
    skipped_existing = 0
    invalid = 0
    issues: list[UserImportIssue] = []
    seen_usernames: set[str] = set()

    for index, row in enumerate(rows, start=2):
        full_name = (row.get("full_name") or row.get("姓名") or "").strip()
        if not full_name:
            invalid += 1
            issues.append(UserImportIssue(row_number=index, full_name=None, login_account=None, reason="missing_full_name"))
            continue

        role = _parse_managed_role(row.get("role") or row.get("身份"))
        if role is None:
            invalid += 1
            issues.append(UserImportIssue(row_number=index, full_name=full_name, login_account=None, reason="invalid_role"))
            continue

        classroom_name = _normalized_classroom_name(row.get("class_name") or row.get("班级"))
        if role == UserRole.TEACHER:
            grade = None
            classroom_name = None
        else:
            raw_grade = (row.get("grade") or row.get("年级") or "").strip()
            try:
                grade = int(raw_grade) if raw_grade else None
            except ValueError:
                invalid += 1
                issues.append(UserImportIssue(row_number=index, full_name=full_name, login_account=None, reason="invalid_grade"))
                continue

        try:
            username = build_generated_username(full_name, role, classroom_name)
            _validate_managed_user_fields(role, grade=grade, classroom_name=classroom_name, is_graduated=False)
        except (HTTPException, ValueError) as exc:
            invalid += 1
            reason = exc.detail if isinstance(exc, HTTPException) else str(exc)
            issues.append(UserImportIssue(row_number=index, full_name=full_name, login_account=None, reason=reason))
            continue

        if username in seen_usernames:
            invalid += 1
            issues.append(UserImportIssue(row_number=index, full_name=full_name, login_account=username, reason="duplicate_login_account_in_file"))
            continue
        seen_usernames.add(username)

        existing = db.scalar(select(User).where(User.username == username).limit(1))
        if existing:
            skipped_existing += 1
            issues.append(UserImportIssue(row_number=index, full_name=full_name, login_account=username, reason="login_account_already_exists"))
            continue

        user = User(
            username=username,
            student_no=None,
            full_name=full_name,
            role=role,
            password_hash=get_password_hash(build_default_password(full_name)),
            must_change_password=role == UserRole.STUDENT,
        )
        _apply_managed_user_payload(
            db,
            user,
            full_name=full_name,
            role=role,
            grade=grade,
            classroom_name=classroom_name,
            is_graduated=False,
        )
        db.add(user)
        created += 1

    db.commit()
    audit_service.log(
        db,
        actor=current_user,
        action="import_users",
        target_type="batch",
        target_id=None,
        result="success",
        ip_address=request.client.host if request and request.client else None,
        detail={"created": created, "rows": len(rows), "skipped_existing": skipped_existing, "invalid": invalid},
    )
    return UserImportResult(rows=len(rows), created=created, skipped_existing=skipped_existing, invalid=invalid, issues=issues[:50])


@router.post("/users/import", response_model=UserImportResult)
async def import_users(
    file: UploadFile = File(...),
    db: DbSession = None,
    current_user: CurrentAdmin = None,
    request: Request = None,
) -> UserImportResult:
    content = await file.read()
    return _import_users(file.filename or "", content, db, current_user, request)


@router.post("/students/import", response_model=UserImportResult)
async def import_students_legacy(
    file: UploadFile = File(...),
    db: DbSession = None,
    current_user: CurrentAdmin = None,
    request: Request = None,
) -> UserImportResult:
    content = await file.read()
    return _import_users(file.filename or "", content, db, current_user, request)


def _conversation_archive_query(
    *,
    student_id: int | None = None,
    subject: str | None = None,
    deleted_by_student: bool | None = None,
):
    query = (
        select(Conversation)
        .options(
            selectinload(Conversation.student).selectinload(User.classroom),
            selectinload(Conversation.messages),
        )
        .order_by(Conversation.updated_at.desc(), Conversation.id.desc())
    )
    if student_id is not None:
        query = query.where(Conversation.student_id == student_id)
    normalized_subject = (subject or "").strip()
    if normalized_subject:
        query = query.where(Conversation.subject == normalized_subject)
    if deleted_by_student is True:
        query = query.where(Conversation.deleted_by_student_at.is_not(None))
    elif deleted_by_student is False:
        query = query.where(Conversation.deleted_by_student_at.is_(None))
    return query


def _archive_message_read(message: Message) -> ConversationArchiveMessageRead:
    return ConversationArchiveMessageRead(
        id=message.id,
        role=message.role,
        content=message.content,
        turn_index=message.turn_index,
        guidance_stage=message.guidance_stage,
        created_at=message.created_at,
    )


def _archive_read(conversation: Conversation) -> ConversationArchiveRead:
    student = conversation.student
    return ConversationArchiveRead(
        id=conversation.id,
        student_id=conversation.student_id,
        student_name=student.full_name if student else "",
        student_username=student.username if student else "",
        student_grade=student.grade if student else None,
        grade_label=student.grade_label if student else None,
        classroom_name=student.classroom_name if student else None,
        classroom_label=student.classroom_label if student else None,
        subject=conversation.subject,
        topic=conversation.topic,
        guidance_stage=conversation.guidance_stage,
        resolved=conversation.resolved,
        duration_seconds=conversation.duration_seconds,
        deleted_by_student=conversation.deleted_by_student_at is not None,
        deleted_by_student_at=conversation.deleted_by_student_at,
        created_at=conversation.created_at,
        updated_at=conversation.updated_at,
        messages=[_archive_message_read(message) for message in conversation.messages],
    )


@router.get("/conversation-archive", response_model=list[ConversationArchiveRead])
def list_conversation_archive(
    db: DbSession,
    current_user: CurrentAdmin,
    student_id: int | None = None,
    subject: str | None = None,
    deleted_by_student: bool | None = None,
    limit: int = 200,
) -> list[ConversationArchiveRead]:
    bounded_limit = max(1, min(limit, 1000))
    rows = db.scalars(
        _conversation_archive_query(
            student_id=student_id,
            subject=subject,
            deleted_by_student=deleted_by_student,
        ).limit(bounded_limit)
    ).all()
    return [_archive_read(row) for row in rows]


@router.get("/conversation-archive/export")
def export_conversation_archive(
    db: DbSession,
    current_user: CurrentAdmin,
    student_id: int | None = None,
    subject: str | None = None,
    deleted_by_student: bool | None = None,
) -> Response:
    rows = db.scalars(
        _conversation_archive_query(
            student_id=student_id,
            subject=subject,
            deleted_by_student=deleted_by_student,
        )
    ).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "conversation_id",
            "student_id",
            "student_name",
            "student_username",
            "grade_label",
            "classroom_label",
            "subject",
            "topic",
            "conversation_created_at",
            "conversation_updated_at",
            "deleted_by_student",
            "deleted_by_student_at",
            "message_id",
            "message_role",
            "turn_index",
            "message_guidance_stage",
            "message_created_at",
            "message_content",
        ]
    )
    for conversation in rows:
        archive = _archive_read(conversation)
        messages = archive.messages or [None]
        for message in messages:
            writer.writerow(
                [
                    archive.id,
                    archive.student_id,
                    archive.student_name,
                    archive.student_username,
                    archive.grade_label or "",
                    archive.classroom_label or "",
                    archive.subject,
                    archive.topic,
                    archive.created_at.isoformat(),
                    archive.updated_at.isoformat(),
                    "true" if archive.deleted_by_student else "false",
                    archive.deleted_by_student_at.isoformat() if archive.deleted_by_student_at else "",
                    message.id if message else "",
                    message.role.value if message else "",
                    message.turn_index if message else "",
                    message.guidance_stage.value if message else "",
                    message.created_at.isoformat() if message else "",
                    message.content if message else "",
                ]
            )
    return Response(
        content="\ufeff" + output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="conversation-archive.csv"'},
    )


@router.get("/audit-logs", response_model=list[AuditLogRead])
def list_audit_logs(
    limit: int = 200,
    action: str | None = None,
    result: str | None = None,
    db: DbSession = None,
    current_user: CurrentAdmin = None,
) -> list[AuditLogRead]:
    query = select(AuditLog).options(selectinload(AuditLog.actor)).order_by(AuditLog.created_at.desc()).limit(limit)
    if action:
        query = query.where(AuditLog.action == action)
    if result:
        query = query.where(AuditLog.result == result)
    rows = db.scalars(query).all()
    return [AuditLogRead.model_validate(row) for row in rows]
