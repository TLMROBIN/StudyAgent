import csv
import io
from typing import Literal

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import PlainTextResponse, Response

from backend.dependencies import CurrentUser, DbSession
from backend.models.schemas import ClassroomStat, StatsOverview, StudentPortrait, StudentProfile
from backend.models.user import UserRole
from backend.services.stats_service import stats_service
from backend.time_utils import serialize_datetime_for_api

router = APIRouter(prefix="/api/stats", tags=["stats"])


@router.get("/overview", response_model=StatsOverview)
def overview(db: DbSession, current_user: CurrentUser) -> StatsOverview:
    return StatsOverview(**stats_service.overview(db, current_user))


@router.get("/students/{student_id}", response_model=StudentProfile)
def student_profile(student_id: int, db: DbSession, current_user: CurrentUser) -> StudentProfile:
    if current_user.role == UserRole.STUDENT and current_user.id != student_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    return StudentProfile(**stats_service.student_profile(db, student_id))


@router.get("/classes", response_model=list[ClassroomStat])
def classroom_breakdown(db: DbSession, current_user: CurrentUser) -> list[ClassroomStat]:
    if current_user.role == UserRole.STUDENT:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    rows = stats_service.classroom_breakdown(db, current_user)
    return [ClassroomStat(**row) for row in rows]


@router.get("/portraits", response_model=list[StudentPortrait])
def student_portraits(limit: int = 12, db: DbSession = None, current_user: CurrentUser = None) -> list[StudentPortrait]:
    if current_user.role == UserRole.STUDENT:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")
    rows = stats_service.student_portraits(db, current_user, limit=limit)
    return [StudentPortrait(**row) for row in rows]


def _build_csv_export(overview: dict) -> PlainTextResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["metric", "value"])
    writer.writerow(["total_questions", overview["total_questions"]])
    writer.writerow(["resolved_rate", overview["resolved_rate"]])
    writer.writerow(["average_turns", overview["average_turns"]])
    for subject_row in overview["by_subject"]:
        writer.writerow([f"subject:{subject_row['subject']}", subject_row["count"]])
    return PlainTextResponse(buffer.getvalue(), media_type="text/csv; charset=utf-8")


def _build_xlsx_export(overview: dict, class_rows: list[dict], portrait_rows: list[dict]) -> Response:
    from openpyxl import Workbook

    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "概览"
    overview_sheet.append(["指标", "值"])
    overview_sheet.append(["累计提问", overview["total_questions"]])
    overview_sheet.append(["已解决率", overview["resolved_rate"]])
    overview_sheet.append(["平均轮次", overview["average_turns"]])

    subject_sheet = workbook.create_sheet("学科分布")
    subject_sheet.append(["学科", "次数"])
    for row in overview["by_subject"]:
        subject_sheet.append([row["subject"], row["count"]])

    class_sheet = workbook.create_sheet("班级统计")
    class_sheet.append(["班级", "年级", "班级名", "学生数", "会话数", "解决率", "平均轮次"])
    for row in class_rows:
        class_sheet.append(
            [
                row["classroom_label"],
                row["grade"],
                row["classroom_name"],
                row["student_count"],
                row["total_conversations"],
                row["resolved_rate"],
                row["average_turns"],
            ]
        )

    portrait_sheet = workbook.create_sheet("学生画像")
    portrait_sheet.append(["学生", "登录账号", "班级", "会话数", "解决率", "关注学科", "兜底占比", "最近活跃"])
    for row in portrait_rows:
        portrait_sheet.append(
            [
                row["student_name"],
                row["login_account"],
                row["classroom_label"],
                row["total_conversations"],
                row["resolved_rate"],
                row["focus_subject"],
                row["fallback_ratio"],
                serialize_datetime_for_api(row["last_active_at"]) if row["last_active_at"] else None,
            ]
        )

    payload = io.BytesIO()
    workbook.save(payload)
    return Response(
        content=payload.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/export")
def export_stats(
    db: DbSession,
    current_user: CurrentUser,
    format: Literal["csv", "xlsx"] = "csv",
) -> Response:
    overview = stats_service.overview(db, current_user)
    if format == "xlsx":
        class_rows = stats_service.classroom_breakdown(db, current_user)
        portrait_rows = stats_service.student_portraits(db, current_user, limit=100)
        return _build_xlsx_export(overview, class_rows, portrait_rows)
    return _build_csv_export(overview)
