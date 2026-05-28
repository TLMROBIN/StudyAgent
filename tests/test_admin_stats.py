import asyncio
import io
from datetime import UTC, datetime, timedelta
from types import SimpleNamespace

from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from backend.database import Base
from backend.database import get_db
from backend.dependencies import get_current_user
from backend.models import agent_config, audit_log, conversation, knowledge, user  # noqa: F401
from backend.models.conversation import Conversation, GuidanceStage, Message, MessageRole
from backend.models.schemas import StudentPortrait
from backend.models.user import Classroom, User, UserRole
from backend.routers import admin as admin_router
from backend.routers import stats as stats_router
from backend.services.account_service import build_default_password, build_generated_username
from backend.services.auth_service import auth_service
from backend.services.stats_service import stats_service


def build_session():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    SessionLocal = sessionmaker(bind=engine, class_=Session, expire_on_commit=False)
    Base.metadata.create_all(bind=engine)
    return SessionLocal


def build_admin_client(session_factory, current_user: User) -> TestClient:
    app = FastAPI()
    app.include_router(admin_router.router)

    def override_db():
        session = session_factory()
        try:
            yield session
        finally:
            session.close()

    def override_current_user():
        return current_user

    app.dependency_overrides[get_db] = override_db
    app.dependency_overrides[get_current_user] = override_current_user
    return TestClient(app)


class FakeUploadFile:
    def __init__(self, filename: str, payload: bytes):
        self.filename = filename
        self._payload = payload

    async def read(self) -> bytes:
        return self._payload


def test_stats_service_returns_classroom_breakdown_and_portraits():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        classroom = Classroom(grade=1, name="1班")
        teacher = User(
            username="mathteacher",
            full_name="数学老师",
            role=UserRole.TEACHER,
            password_hash="hash",
        )
        teacher.teacher_classrooms.append(classroom)
        student = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=classroom,
        )
        session.add_all([classroom, teacher, student])
        session.commit()
        session.refresh(student)

        conversation_row = Conversation(
            student_id=student.id,
            subject="数学",
            resolved=True,
            guidance_stage=GuidanceStage.FALLBACK,
        )
        session.add(conversation_row)
        session.commit()
        session.refresh(conversation_row)
        session.add_all(
            [
                Message(conversation_id=conversation_row.id, role=MessageRole.USER, content="函数单调性怎么判断"),
                Message(conversation_id=conversation_row.id, role=MessageRole.ASSISTANT, content="先看定义域"),
            ]
        )
        session.commit()
        session.refresh(teacher)

        classroom_rows = stats_service.classroom_breakdown(session, teacher)
        portrait_rows = stats_service.student_portraits(session, teacher)

        assert len(classroom_rows) == 1
        assert classroom_rows[0]["classroom_label"] == "高一1班"
        assert classroom_rows[0]["student_count"] == 1
        assert classroom_rows[0]["total_conversations"] == 1

        assert len(portrait_rows) == 1
        assert portrait_rows[0]["student_name"] == "张三"
        assert portrait_rows[0]["login_account"] == "zhangsan1"
        assert portrait_rows[0]["focus_subject"] == "数学"
        assert portrait_rows[0]["fallback_ratio"] == 1.0
    finally:
        session.close()


def test_stats_service_usage_trend_groups_by_day_and_selected_subjects():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        classroom = Classroom(grade=1, name="1班")
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        student = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=classroom,
        )
        session.add_all([classroom, admin_user, student])
        session.commit()
        session.refresh(student)

        rows = [
            Conversation(
                student_id=student.id,
                subject="数学",
                created_at=datetime(2026, 5, 1, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 5, 1, 9, 0, tzinfo=UTC),
            ),
            Conversation(
                student_id=student.id,
                subject="物理",
                created_at=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
                updated_at=datetime(2026, 5, 1, 10, 0, tzinfo=UTC),
            ),
            Conversation(
                student_id=student.id,
                subject="数学",
                created_at=datetime(2026, 5, 3, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 5, 3, 9, 0, tzinfo=UTC),
            ),
        ]
        session.add_all(rows)
        session.commit()

        trend = stats_service.usage_trend(
            session,
            admin_user,
            granularity="day",
            start_date="2026-05-01",
            end_date="2026-05-03",
            subjects=["数学"],
        )

        assert trend["granularity"] == "day"
        assert trend["labels"] == ["2026-05-01", "2026-05-02", "2026-05-03"]
        assert trend["available_subjects"] == ["数学", "物理"]
        assert trend["series"] == [
            {"name": "总次数", "subject": None, "data": [2, 0, 1]},
            {"name": "数学", "subject": "数学", "data": [1, 0, 1]},
        ]
    finally:
        session.close()


def test_import_users_returns_detailed_feedback(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        monkeypatch.setattr(admin_router, "get_password_hash", lambda password: f"hash:{password}")
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        existing_student = User(
            username=build_generated_username("张三", UserRole.STUDENT, "1班"),
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=Classroom(grade=1, name="1班"),
        )
        session.add_all([admin_user, existing_student])
        session.commit()
        session.refresh(admin_user)

        csv_payload = (
            "full_name,role,grade,class_name\n"
            "李四,学生,1,1班\n"
            "张三,学生,1,1班\n"
            ",学生,1,1班\n"
            "王五,学生,abc,1班\n"
            "李四,学生,1,1班\n"
        )
        upload = FakeUploadFile(filename="users.csv", payload=csv_payload.encode("utf-8"))
        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))

        result = asyncio.run(
            admin_router.import_users(
                file=upload,
                db=session,
                current_user=admin_user,
                request=request,
            )
        )

        assert result.rows == 5
        assert result.created == 1
        assert result.skipped_existing == 1
        assert result.invalid == 3
        assert {issue.reason for issue in result.issues} >= {
            "login_account_already_exists",
            "missing_full_name",
            "invalid_grade",
            "duplicate_login_account_in_file",
        }
    finally:
        session.close()


def test_import_users_supports_xlsx_and_teacher_rows(monkeypatch):
    from openpyxl import Workbook

    SessionLocal = build_session()
    session = SessionLocal()
    try:
        monkeypatch.setattr(admin_router, "get_password_hash", lambda password: f"hash:{password}")
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        session.add(admin_user)
        session.commit()
        session.refresh(admin_user)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["full_name", "role", "grade", "class_name"])
        sheet.append(["表格学生", "student", 2, "3班"])
        sheet.append(["赵老师", "teacher", None, None])
        payload = io.BytesIO()
        workbook.save(payload)

        upload = FakeUploadFile(filename="users.xlsx", payload=payload.getvalue())
        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))

        result = asyncio.run(
            admin_router.import_users(
                file=upload,
                db=session,
                current_user=admin_user,
                request=request,
            )
        )

        created_student = session.query(User).filter(User.full_name == "表格学生").one()
        created_teacher = session.query(User).filter(User.full_name == "赵老师").one()
        assert result.rows == 2
        assert result.created == 2
        assert created_student.username == build_generated_username("表格学生", UserRole.STUDENT, "3班")
        assert created_student.grade_label == "高二"
        assert created_teacher.username == build_generated_username("赵老师", UserRole.TEACHER)
        assert created_teacher.classroom_id is None
        assert created_teacher.grade is None
    finally:
        session.close()


def test_list_users_supports_grade_classroom_and_name_filters():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        class_one = Classroom(grade=1, name="1班")
        class_two = Classroom(grade=2, name="2班")
        zhang = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=class_one,
        )
        li = User(
            username="lisi2",
            full_name="李四",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=2,
            classroom=class_two,
        )
        teacher = User(
            username="wanglaoshi",
            full_name="王老师",
            role=UserRole.TEACHER,
            password_hash="hash",
        )
        session.add_all([admin_user, class_one, class_two, zhang, li, teacher])
        session.commit()
        session.refresh(admin_user)

        filtered = admin_router.list_users(
            session,
            admin_user,
            grade="1",
            classroom_name="1班",
            keyword="张",
        )
        teacher_match = admin_router.list_users(session, admin_user, keyword="王")
        empty = admin_router.list_users(session, admin_user, grade="1", classroom_name="2班")

        assert [item.full_name for item in filtered] == ["张三"]
        assert [item.full_name for item in teacher_match] == ["王老师"]
        assert empty == []
    finally:
        session.close()


def test_list_classroom_options_can_follow_grade_filter():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        class_one = Classroom(grade=1, name="1班")
        class_two = Classroom(grade=2, name="2班")
        class_three = Classroom(grade=2, name="3班")
        session.add_all([admin_user, class_one, class_two, class_three])
        session.commit()
        session.refresh(admin_user)

        all_options = admin_router.list_classrooms(session, admin_user)
        grade_two_options = admin_router.list_classrooms(session, admin_user, grade="2")
        unset_options = admin_router.list_classrooms(session, admin_user, grade="unset")

        assert [(item.grade, item.name) for item in all_options] == [(1, "1班"), (2, "2班"), (2, "3班")]
        assert [(item.grade, item.name) for item in grade_two_options] == [(2, "2班"), (2, "3班")]
        assert unset_options == []
    finally:
        session.close()


def test_create_user_auto_generates_username_and_default_password(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        monkeypatch.setattr(admin_router, "get_password_hash", lambda password: f"hash:{password}")
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        session.add(admin_user)
        session.commit()
        session.refresh(admin_user)

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        result = admin_router.create_user(
            admin_router.UserCreate(
                full_name="高二学生",
                role=UserRole.STUDENT,
                grade=2,
                classroom_name="2班",
            ),
            db=session,
            current_user=admin_user,
            request=request,
        )

        created_student = session.query(User).filter(User.full_name == "高二学生").one()
        assert result.username == build_generated_username("高二学生", UserRole.STUDENT, "2班")
        assert result.grade_label == "高二"
        assert created_student.password_hash == f"hash:{build_default_password('高二学生')}"
        assert created_student.classroom_label == "高二2班"
    finally:
        session.close()


def test_update_user_can_switch_student_to_teacher(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        student = User(
            username=build_generated_username("待调整", UserRole.STUDENT, "3班"),
            full_name="待调整",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=3,
            classroom=Classroom(grade=3, name="3班"),
        )
        session.add_all([admin_user, student])
        session.commit()
        session.refresh(admin_user)
        session.refresh(student)

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        result = admin_router.update_user(
            student.id,
            admin_router.UserUpdate(
                full_name="赵老师",
                role=UserRole.TEACHER,
                grade=None,
                classroom_name=None,
                is_graduated=False,
                is_active=False,
            ),
            db=session,
            current_user=admin_user,
            request=request,
        )

        refreshed = session.get(User, student.id)
        assert refreshed is not None
        assert refreshed.role == UserRole.TEACHER
        assert refreshed.grade is None
        assert refreshed.classroom_id is None
        assert refreshed.username == build_generated_username("赵老师", UserRole.TEACHER)
        assert result.is_active is False
    finally:
        session.close()


def test_delete_user_removes_teacher(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        teacher = User(
            username="zhaolaoshi",
            full_name="赵老师",
            role=UserRole.TEACHER,
            password_hash="hash",
        )
        session.add_all([admin_user, teacher])
        session.commit()
        session.refresh(admin_user)

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        admin_router.delete_user(teacher.id, session, admin_user, request)

        assert session.get(User, teacher.id) is None
    finally:
        session.close()


def test_delete_user_rejects_student_with_archived_conversations(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        student = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=Classroom(grade=1, name="1班"),
        )
        session.add_all([admin_user, student])
        session.commit()
        session.refresh(admin_user)
        session.refresh(student)

        conversation_row = Conversation(
            student_id=student.id,
            subject="数学",
            guidance_stage=GuidanceStage.INITIAL,
        )
        session.add(conversation_row)
        session.commit()
        session.refresh(conversation_row)
        session.add(Message(conversation_id=conversation_row.id, role=MessageRole.USER, content="函数怎么做"))
        session.commit()

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        try:
            admin_router.delete_user(student.id, session, admin_user, request)
        except Exception as exc:
            assert getattr(exc, "status_code", None) == 400
            assert "archived conversations" in str(getattr(exc, "detail", ""))
        else:
            raise AssertionError("Expected student deletion to be rejected when conversations exist")

        assert session.get(User, student.id) is not None
        assert session.get(Conversation, conversation_row.id) is not None
    finally:
        session.close()


def test_reset_password_uses_generated_default(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        monkeypatch.setattr(admin_router.auth_service, "update_password", lambda db, user, new_password: setattr(user, "password_hash", f"hash:{new_password}") or user)
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        teacher = User(
            username="zhaolaoshi",
            full_name="赵老师",
            role=UserRole.TEACHER,
            password_hash="old-hash",
        )
        session.add_all([admin_user, teacher])
        session.commit()
        session.refresh(admin_user)
        session.refresh(teacher)

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        result = admin_router.reset_password(
            admin_router.PasswordResetRequest(user_id=teacher.id),
            session,
            admin_user,
            request,
        )

        assert result.username == "zhaolaoshi"
        assert teacher.password_hash == f"hash:{build_default_password('赵老师')}"
    finally:
        session.close()


def test_reset_password_clears_failed_login_lock():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        student = User(
            username="zhoujundong14",
            full_name="周俊栋",
            role=UserRole.STUDENT,
            password_hash="old-hash",
            failed_login_count=6,
            locked_until=datetime.now(UTC) + timedelta(minutes=15),
        )
        session.add_all([admin_user, student])
        session.commit()
        session.refresh(admin_user)
        session.refresh(student)

        request = SimpleNamespace(client=SimpleNamespace(host="127.0.0.1"))
        result = admin_router.reset_password(
            admin_router.PasswordResetRequest(user_id=student.id),
            session,
            admin_user,
            request,
        )

        assert result.username == "zhoujundong14"
        assert student.failed_login_count == 0
        assert student.locked_until is None
    finally:
        session.close()


def test_admin_can_view_student_cleared_conversation_archive():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(username="admin", full_name="管理员", role=UserRole.ADMIN, password_hash="hash")
        student = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=Classroom(grade=1, name="1班"),
        )
        session.add_all([admin_user, student])
        session.commit()
        session.refresh(admin_user)
        session.refresh(student)
        cleared_at = datetime(2026, 5, 23, 9, 30, tzinfo=UTC)
        conversation_row = Conversation(
            student_id=student.id,
            subject="数学",
            guidance_stage=GuidanceStage.HINT,
            deleted_by_student_at=cleared_at,
        )
        session.add(conversation_row)
        session.commit()
        session.refresh(conversation_row)
        session.add_all(
            [
                Message(
                    conversation_id=conversation_row.id,
                    role=MessageRole.USER,
                    content="函数单调性怎么判断",
                    turn_index=1,
                    guidance_stage=GuidanceStage.INITIAL,
                    created_at=datetime(2026, 5, 23, 9, 0, tzinfo=UTC),
                ),
                Message(
                    conversation_id=conversation_row.id,
                    role=MessageRole.ASSISTANT,
                    content="先看定义域。",
                    turn_index=1,
                    guidance_stage=GuidanceStage.HINT,
                    llm_model_key="minimax-m27",
                    created_at=datetime(2026, 5, 23, 9, 1, tzinfo=UTC),
                ),
            ]
        )
        session.commit()
    finally:
        session.close()

    client = build_admin_client(SessionLocal, admin_user)
    response = client.get("/api/admin/conversation-archive")

    assert response.status_code == 200
    payload = response.json()
    assert payload[0]["student_name"] == "张三"
    assert payload[0]["student_username"] == "zhangsan1"
    assert payload[0]["classroom_label"] == "高一1班"
    assert payload[0]["subject"] == "数学"
    assert payload[0]["deleted_by_student"] is True
    assert payload[0]["deleted_by_student_at"] is not None
    assert [message["content"] for message in payload[0]["messages"]] == ["函数单调性怎么判断", "先看定义域。"]
    assert payload[0]["messages"][0]["llm_model_key"] is None
    assert payload[0]["messages"][1]["llm_model_key"] == "minimax-m27"


def test_admin_can_export_conversation_archive_csv():
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        admin_user = User(username="admin", full_name="管理员", role=UserRole.ADMIN, password_hash="hash")
        student = User(
            username="lisi1",
            full_name="李四",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=2,
            classroom=Classroom(grade=2, name="3班"),
        )
        session.add_all([admin_user, student])
        session.commit()
        session.refresh(admin_user)
        session.refresh(student)
        conversation_row = Conversation(student_id=student.id, subject="物理")
        session.add(conversation_row)
        session.commit()
        session.refresh(conversation_row)
        session.add_all(
            [
                Message(conversation_id=conversation_row.id, role=MessageRole.USER, content="牛顿第二定律怎么用", turn_index=1),
                Message(
                    conversation_id=conversation_row.id,
                    role=MessageRole.ASSISTANT,
                    content="先判断受力。",
                    turn_index=1,
                    llm_model_key="qwen2.5-vl",
                ),
            ]
        )
        session.commit()
    finally:
        session.close()

    client = build_admin_client(SessionLocal, admin_user)
    response = client.get("/api/admin/conversation-archive/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    csv_text = response.content.decode("utf-8-sig")
    assert "conversation_id,student_id,student_name" in csv_text
    assert "message_llm_model_key" in csv_text
    assert "李四" in csv_text
    assert "牛顿第二定律怎么用" in csv_text
    assert "qwen2.5-vl" in csv_text


def test_student_can_authenticate_with_generated_login_account(monkeypatch):
    SessionLocal = build_session()
    session = SessionLocal()
    try:
        student = User(
            username=build_generated_username("张三", UserRole.STUDENT, "1班"),
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash=admin_router.get_password_hash(build_default_password("张三")),
            grade=1,
            classroom=Classroom(grade=1, name="1班"),
        )
        session.add(student)
        session.commit()

        authenticated = auth_service.authenticate_student(session, "zhangsan1", build_default_password("张三"))

        assert authenticated is not None
        assert authenticated.username == "zhangsan1"
    finally:
        session.close()


def test_stats_export_supports_xlsx():
    from openpyxl import load_workbook

    SessionLocal = build_session()
    session = SessionLocal()
    try:
        classroom = Classroom(grade=1, name="1班")
        admin_user = User(
            username="admin",
            full_name="管理员",
            role=UserRole.ADMIN,
            password_hash="hash",
        )
        student = User(
            username="zhangsan1",
            full_name="张三",
            role=UserRole.STUDENT,
            password_hash="hash",
            grade=1,
            classroom=classroom,
        )
        session.add_all([classroom, admin_user, student])
        session.commit()

        conversation_row = Conversation(
            student_id=student.id,
            subject="数学",
            resolved=True,
            guidance_stage=GuidanceStage.HINT,
        )
        session.add(conversation_row)
        session.commit()
        session.add_all(
            [
                Message(conversation_id=conversation_row.id, role=MessageRole.USER, content="函数题怎么做"),
                Message(conversation_id=conversation_row.id, role=MessageRole.ASSISTANT, content="先看定义域"),
            ]
        )
        session.commit()

        response = stats_router.export_stats(session, admin_user, format="xlsx")
        workbook = load_workbook(io.BytesIO(response.body))

        assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert workbook.sheetnames == ["概览", "学科分布", "班级统计", "学生画像"]
        assert workbook["概览"]["A2"].value == "累计提问"
        assert workbook["概览"]["B2"].value == 1
        assert workbook["学科分布"]["A2"].value == "数学"
        assert workbook["学生画像"]["B2"].value == "zhangsan1"
        assert str(workbook["学生画像"]["H2"].value).endswith("+08:00")
    finally:
        session.close()


def test_student_portrait_serializes_naive_utc_to_beijing():
    portrait = StudentPortrait(
        student_id=1,
        student_name="张三",
        login_account="zhangsan1",
        total_conversations=1,
        resolved_rate=1.0,
        fallback_ratio=0.0,
        last_active_at=datetime(2026, 4, 9, 10, 15, 0),
    )

    payload = portrait.model_dump(mode="json")

    assert payload["last_active_at"] == "2026-04-09T18:15:00+08:00"
